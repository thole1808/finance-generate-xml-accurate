from __future__ import annotations

import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from xml.dom import minidom
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
import xlrd


SUPPORTED_EXCEL_EXTENSIONS = {".xls", ".xlsx", ".xlsm", ".xltx", ".xltm"}
MODERN_EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
LEGACY_EXCEL_EXTENSIONS = {".xls"}


def get_sheet_names(excel_path: str | Path) -> list[str]:
    path = Path(excel_path)
    _validate_excel_path(path)
    if path.suffix.lower() in LEGACY_EXCEL_EXTENSIONS:
        workbook = xlrd.open_workbook(path)
        return workbook.sheet_names()

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def read_excel_rows(
    excel_path: str | Path,
    sheet_name: str | None = None,
    max_rows: int | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    path = Path(excel_path)
    _validate_excel_path(path)

    if path.suffix.lower() in LEGACY_EXCEL_EXTENSIONS:
        return _read_xls_rows(path, sheet_name=sheet_name, max_rows=max_rows)

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)

        try:
            raw_headers = next(rows)
        except StopIteration:
            return [], []

        headers = _normalize_headers(raw_headers)
        records: list[dict[str, Any]] = []

        for index, row in enumerate(rows, start=1):
            if max_rows is not None and index > max_rows:
                break

            if not row or all(value is None or str(value).strip() == "" for value in row):
                continue

            record = {
                headers[column_index]: value
                for column_index, value in enumerate(row[: len(headers)])
                if column_index < len(headers)
            }
            records.append(record)

        return headers, records
    finally:
        workbook.close()


def excel_to_accurate_xml(
    excel_path: str | Path,
    output_path: str | Path,
    sheet_name: str | None = None,
    document_type: str = "GENERAL",
) -> int:
    headers, rows = read_excel_rows(excel_path, sheet_name=sheet_name)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    if document_type == "SALES_INVOICE":
        output.write_text(
            _build_sales_invoice_xml(rows, branch_code="2012585710", exim_id="27"),
            encoding="utf-8",
        )
        return len(rows)
    if document_type == "SALES_RECEIPT":
        output.write_text(
            _build_sales_receipt_xml(rows, branch_code="2012585710", exim_id="27"),
            encoding="utf-8",
        )
        return len(rows)
    if document_type == "OTHER_PAYMENT":
        output.write_text(
            _build_other_payment_xml(rows, branch_code="2012585710", exim_id="27"),
            encoding="utf-8",
        )
        return len(rows)

    root = ET.Element("ACCURATE_IMPORT")
    metadata = ET.SubElement(root, "METADATA")
    ET.SubElement(metadata, "DOCUMENT_TYPE").text = document_type
    ET.SubElement(metadata, "SOURCE_FILE").text = Path(excel_path).name
    ET.SubElement(metadata, "TOTAL_ROWS").text = str(len(rows))

    data = ET.SubElement(root, "DATA")
    for number, row in enumerate(rows, start=1):
        row_element = ET.SubElement(data, "ROW", number=str(number))
        for header in headers:
            tag_name = _xml_tag_name(header)
            value = row.get(header)
            ET.SubElement(row_element, tag_name).text = "" if value is None else str(value)

    output.write_text(_prettify_xml(root), encoding="utf-8")
    return len(rows)


def _build_sales_invoice_xml(
    rows: list[dict[str, Any]],
    branch_code: str,
    exim_id: str,
) -> str:
    root = ET.Element(
        "NMEXML",
        {
            "ACCOUNTANTCOPYID": "",
            "BranchCode": branch_code,
            "EximID": exim_id,
        },
    )
    transactions = ET.SubElement(root, "TRANSACTIONS", {"OnError": "CONTINUE"})

    for row in rows:
        invoice = ET.SubElement(transactions, "SALESINVOICE", {"REQUESTID": "1", "operation": "Add"})
        _add_text(invoice, "TRANSACTIONID", _field(row, "TRANSACTIONID", default="880"))
        _add_sales_invoice_itemline(invoice, row)

        _add_text(invoice, "INVOICENO", _field(row, "INVOICENO"))
        invoice_date = _date_field(row, "INVOICEDATE")
        _add_text(invoice, "INVOICEDATE", invoice_date)
        _add_text(invoice, "TAX1ID", _field(row, "TAX1ID"))
        _add_text(invoice, "TAX1CODE", _field(row, "TAX1CODE"))
        _add_text(invoice, "TAX2CODE", _field(row, "TAX2CODE"))
        _add_text(invoice, "TAX1RATE", _field(row, "TAX1RATE"))
        _add_text(invoice, "TAX2RATE", _field(row, "TAX2RATE", default="0"))
        _add_text(invoice, "RATE", _field(row, "RATE", default="1"))
        _add_text(invoice, "INCLUSIVETAX", _field(row, "INCLUSIVETAX", default="0"))
        _add_text(invoice, "CUSTOMERISTAXABLE", _field(row, "CUSTOMERISTAXABLE", default="1"))
        _add_text(invoice, "CASHDISCOUNT", _field(row, "CASHDISCOUNT"))
        _add_text(invoice, "CASHDISCPC", _field(row, "CASHDISCPC"))
        _add_text(invoice, "INVOICEAMOUNT", _field(row, "INVOICEAMOUNT", default="0"))
        _add_text(invoice, "FREIGHT", _field(row, "FREIGHT", default="0"))
        _add_text(invoice, "TERMSID", _field(row, "TERMSID", default="C.O.D"))
        _add_text(invoice, "FOB", _field(row, "FOB"))
        _add_text(invoice, "PURCHASEORDERNO", _field(row, "PURCHASEORDERNO"))
        _add_text(invoice, "WAREHOUSEID", _field(row, "WAREHOUSEID"))
        _add_text(invoice, "DESCRIPTION", _field(row, "DESCRIPTION"))
        _add_text(invoice, "SHIPDATE", _date_field(row, "SHIPDATE", fallback=invoice_date))
        _add_text(invoice, "DELIVERYORDER", _field(row, "DELIVERYORDER"))
        _add_text(invoice, "FISCALRATE", _field(row, "FISCALRATE", default="1"))
        _add_text(invoice, "TAXDATE", _date_field(row, "TAXDATE", fallback=invoice_date))
        _add_text(invoice, "CUSTOMERID", _field(row, "CUSTOMERID"))

        salesman = ET.SubElement(invoice, "SALESMANID")
        _add_text(salesman, "LASTNAME", _field(row, "SALESMAN_LASTNAME"))
        _add_text(salesman, "FIRSTNAME", _field(row, "SALESMAN", "SALESMAN_FIRSTNAME"))

        _add_text(invoice, "PRINTED", _field(row, "PRINTED", default="0"))
        for index in range(1, 6):
            _add_text(invoice, f"SHIPTO{index}", _field(row, f"SHIPTO{index}"))
        _add_text(invoice, "ARACCOUNT", _field(row, "ARACCOUNT"))
        _add_text(invoice, "TAXFORMNUMBER", _field(row, "TAXFORMNUMBER"))
        _add_text(invoice, "TAXFORMCODE", _field(row, "TAXFORMCODE"))
        _add_text(invoice, "CURRENCYNAME", _field(row, "CURRENCYNAME", default="IDR"))
        _add_text(invoice, "AUTOMATICINSERTGROUPING", _field(row, "AUTOMATICINSERTGROUPING"))

    return _prettify_xml(root)


def _build_sales_receipt_xml(
    rows: list[dict[str, Any]],
    branch_code: str,
    exim_id: str,
) -> str:
    root = _accurate_root(branch_code, exim_id)
    transactions = ET.SubElement(root, "TRANSACTIONS", {"OnError": "CONTINUE"})

    for row in rows:
        receipt = ET.SubElement(transactions, "CUSTOMERRECEIPT", {"REQUESTID": "1", "operation": "Add"})
        _add_text(receipt, "TRANSACTIONID", _field(row, "TRANSACTIONID", default="617"))
        _add_text(receipt, "IMPORTEDTRANSACTIONID", _field(row, "IMPORTEDTRANSACTIONID"))

        invoice_line = ET.SubElement(receipt, "InvoiceLine", {"operation": "Add"})
        _add_text(invoice_line, "KeyID", _field(row, "KeyID", "KEYID", default="1"))
        payment_amount = _field(row, "PAYMENTAMOUNT", "CHEQUEAMOUNT", default="0")
        _add_text(invoice_line, "PAYMENTAMOUNT", payment_amount)
        _add_text(invoice_line, "PPH23AMOUNT", _field(row, "PPH23AMOUNT", default="0"))
        _add_text(invoice_line, "PPH23RATE", _field(row, "PPH23RATE", default="0"))
        _add_text(invoice_line, "PPH23FISCALRATE", _field(row, "PPH23FISCALRATE", default="0"))
        _add_text(invoice_line, "PPH23NUMBER", _field(row, "PPH23NUMBER"))
        _add_text(invoice_line, "DISCTAKENAMOUNT", _field(row, "DISCTAKENAMOUNT", "DI1DISCAMOUNT", default="0"))
        _add_text(invoice_line, "DEPTID", _field(row, "DEPTID"))
        _add_text(invoice_line, "PROJECTID", _field(row, "PROJECTID"))
        _add_text(invoice_line, "ARINVOICEID", _field(row, "ARINVOICEID"))

        payment_date = _date_field(row, "PAYMENTDATE")
        _add_text(receipt, "SEQUENCENO", _field(row, "SEQUENCENO", "CRNO"))
        _add_text(receipt, "PAYMENTDATE", payment_date)
        _add_text(receipt, "CHEQUENO", _field(row, "CHEQUENO"))
        _add_text(receipt, "BANKACCOUNT", _field(row, "BANKACCOUNT"))
        _add_text(receipt, "CHEQUEDATE", _date_field(row, "CHEQUEDATE", fallback=payment_date))
        _add_text(receipt, "CHEQUEAMOUNT", _field(row, "CHEQUEAMOUNT", "PAYMENTAMOUNT", default="0"))
        _add_text(receipt, "RATE", _field(row, "RATE", default="1"))
        _add_text(receipt, "DESCRIPTION", _field(row, "DESCRIPTION"))
        _add_text(receipt, "FISCALPMT", _field(row, "FISCALPMT", default="0"))
        _add_text(receipt, "VOID", _field(row, "VOID", default="0"))
        _add_text(receipt, "BILLTOID", _field(row, "BILLTOID"))
        _add_text(receipt, "OVERPAYUSED", _field(row, "OVERPAYUSED"))
        _add_text(receipt, "APPLYFROMCREDIT", _field(row, "APPLYFROMCREDIT", default="0"))
        _add_text(receipt, "CURRENCYNAME", _field(row, "CURRENCYNAME", default="IDR"))
        _add_text(receipt, "RETURNCREDIT", _field(row, "RETURNCREDIT", default="0"))

    return _prettify_xml(root)


def _build_other_payment_xml(
    rows: list[dict[str, Any]],
    branch_code: str,
    exim_id: str,
) -> str:
    root = _accurate_root(branch_code, exim_id)
    transactions = ET.SubElement(root, "TRANSACTIONS", {"OnError": "CONTINUE"})

    for row in rows:
        payment = ET.SubElement(transactions, "OTHERPAYMENT", {"REQUESTID": "1", "operation": "Add"})
        _add_text(payment, "TRANSACTIONID", _field(row, "TRANSACTIONID", default="853"))

        account_line = ET.SubElement(payment, "ACCOUNTLINE", {"operation": "Add"})
        _add_text(account_line, "KeyID", _field(row, "KeyID", "KEYID", default="1"))
        _add_text(account_line, "GLACCOUNT", _field(row, "GLACCOUNT", "ACCOUNT"))
        amount = _field(row, "GLAMOUNT", "AMOUNT", "JVAMOUNT", default="0")
        _add_text(account_line, "GLAMOUNT", amount)
        _add_text(account_line, "DEPTID", _field(row, "DEPTID"))
        _add_text(account_line, "PROJECTID", _field(row, "PROJECTID"))
        _add_text(account_line, "DESCRIPTION", _field(row, "DESCRIPTION", "TRANSDESCRIPTION"))
        _add_text(account_line, "RATE", _field(row, "RATE", default="1"))
        _add_text(account_line, "PRIMEAMOUNT", _field(row, "PRIMEAMOUNT", "AMOUNT", "JVAMOUNT", default=amount))
        _add_text(account_line, "TXDATE", _date_field(row, "TXDATE"))
        _add_text(account_line, "POSTED", _field(row, "POSTED"))
        _add_text(account_line, "CURRENCYNAME", _field(row, "CURRENCYNAME"))

        trans_date = _date_field(row, "TRANSDATE", "DATE")
        _add_text(payment, "JVNUMBER", _field(row, "JVNUMBER", "OPNO"))
        _add_text(payment, "TRANSDATE", trans_date)
        _add_text(payment, "SOURCE", _field(row, "SOURCE", default="GL"))
        _add_text(payment, "TRANSTYPE", _field(row, "TRANSTYPE", default="other payment"))
        _add_text(payment, "TRANSDESCRIPTION", _field(row, "TRANSDESCRIPTION", "DESCRIPTION"))
        _add_text(payment, "JVAMOUNT", _field(row, "JVAMOUNT", "AMOUNT", default=amount))
        _add_text(payment, "CHEQUENO", _field(row, "CHEQUENO"))
        _add_text(payment, "PAYEE", _field(row, "PAYEE"))
        _add_text(payment, "VOIDCHEQUE", _field(row, "VOIDCHEQUE", default="0"))
        _add_text(payment, "GLACCOUNT", _field(row, "BANKACCOUNT", "PAYMENTACCOUNT"))
        _add_text(payment, "RATE", _field(row, "RATE", default="1"))

    return _prettify_xml(root)


def _accurate_root(branch_code: str, exim_id: str) -> ET.Element:
    return ET.Element(
        "NMEXML",
        {
            "ACCOUNTANTCOPYID": "",
            "BranchCode": branch_code,
            "EximID": exim_id,
        },
    )


def _add_sales_invoice_itemline(parent: ET.Element, row: dict[str, Any]) -> None:
    itemline = ET.SubElement(parent, "ITEMLINE", {"operation": "Add"})
    _add_text(itemline, "KeyID", _field(row, "KeyID", "KEYID", default="1"))
    _add_text(itemline, "ITEMNO", _field(row, "ITEMNO"))
    _add_text(itemline, "QUANTITY", _field(row, "QUANTITY", default="1"))
    _add_text(itemline, "ITEMUNIT", _field(row, "ITEMUNIT"))
    _add_text(itemline, "UNITRATIO", _field(row, "UNITRATIO", default="1"))
    for index in range(1, 11):
        _add_text(itemline, f"ITEMRESERVED{index}", _field(row, f"ITEMRESERVED{index}"))
    _add_text(itemline, "ITEMOVDESC", _field(row, "ITEMOVDESC"))
    unit_price = _field(row, "UNITPRICE")
    _add_text(itemline, "UNITPRICE", unit_price)
    _add_text(itemline, "ITEMDISCPC", _field(row, "ITEMDISCPC"))
    _add_text(itemline, "TAXCODES", _field(row, "TAXCODES"))
    _add_text(itemline, "PROJECTID", _field(row, "PROJECTID"))
    _add_text(itemline, "DEPTID", _field(row, "DEPTID"))
    _add_text(itemline, "GROUPSEQ", _field(row, "GROUPSEQ"))
    _add_text(itemline, "SOSEQ", _field(row, "SOSEQ"))
    _add_text(itemline, "BRUTOUNITPRICE", _field(row, "BRUTOUNITPRICE", default=unit_price))
    _add_text(itemline, "WAREHOUSEID", _field(row, "WAREHOUSEID2", "ITEMWAREHOUSEID", "WAREHOUSEID"))
    _add_text(itemline, "QTYCONTROL", _field(row, "QTYCONTROL", default="0"))
    _add_text(itemline, "DOSEQ", _field(row, "DOSEQ"))
    _add_text(itemline, "DOID", _field(row, "DOID"))


def _add_text(parent: ET.Element, tag: str, value: Any = "") -> ET.Element:
    element = ET.SubElement(parent, tag)
    text = _clean_value(value)
    if text != "":
        element.text = text
    return element


def _field(row: dict[str, Any], *names: str, default: Any = "") -> str:
    normalized = {_normalize_key(key): value for key, value in row.items()}
    for name in names:
        value = normalized.get(_normalize_key(name))
        if value is not None and str(value).strip() != "":
            return _clean_value(value)
    return _clean_value(default)


def _date_field(row: dict[str, Any], *names: str, fallback: Any = "") -> str:
    value = _field(row, *names)
    if not value:
        return _clean_value(fallback)
    return _format_excel_date(value)


def _format_excel_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    text = _clean_value(value)
    if re.fullmatch(r"\d+(\.0)?", text):
        serial = int(float(text))
        return (datetime(1899, 12, 30) + timedelta(days=serial)).strftime("%Y-%m-%d")

    for date_format in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, date_format).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return text


def _clean_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_key(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", value.upper())


def _validate_excel_path(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"File tidak ditemukan: {path}")
    if path.suffix.lower() not in SUPPORTED_EXCEL_EXTENSIONS:
        raise ValueError("Format file harus .xls, .xlsx, .xlsm, .xltx, atau .xltm")


def _read_xls_rows(
    excel_path: Path,
    sheet_name: str | None = None,
    max_rows: int | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = xlrd.open_workbook(excel_path)
    worksheet = workbook.sheet_by_name(sheet_name) if sheet_name else workbook.sheet_by_index(0)

    if worksheet.nrows == 0:
        return [], []

    raw_headers = tuple(_xls_cell_value(worksheet.cell_value(0, column)) for column in range(worksheet.ncols))
    headers = _normalize_headers(raw_headers)
    records: list[dict[str, Any]] = []

    for row_index in range(1, worksheet.nrows):
        if max_rows is not None and len(records) >= max_rows:
            break

        raw_row = tuple(_xls_cell_value(worksheet.cell_value(row_index, column)) for column in range(worksheet.ncols))
        if not raw_row or all(value is None or str(value).strip() == "" for value in raw_row):
            continue

        record = {
            headers[column_index]: value
            for column_index, value in enumerate(raw_row[: len(headers)])
            if column_index < len(headers)
        }
        records.append(record)

    return headers, records


def _xls_cell_value(value: Any) -> Any:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _normalize_headers(raw_headers: tuple[Any, ...]) -> list[str]:
    headers: list[str] = []
    used: dict[str, int] = {}

    for index, value in enumerate(raw_headers, start=1):
        base = str(value).strip() if value is not None and str(value).strip() else f"Column_{index}"
        base = re.sub(r"\s+", "_", base)

        count = used.get(base, 0)
        used[base] = count + 1
        headers.append(base if count == 0 else f"{base}_{count + 1}")

    return headers


def _xml_tag_name(header: str) -> str:
    tag = re.sub(r"[^A-Za-z0-9_.-]", "_", header.strip())
    if not tag:
        return "FIELD"
    if not re.match(r"^[A-Za-z_]", tag):
        return f"FIELD_{tag}"
    return tag


def _prettify_xml(element: ET.Element) -> str:
    rough = ET.tostring(element, encoding="utf-8")
    parsed = minidom.parseString(rough)
    return parsed.toprettyxml(indent="  ", encoding="utf-8").decode("utf-8")
